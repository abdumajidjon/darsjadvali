"""Excel schedule parsing engine with merged cells support"""

import io
from typing import Any
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.merge import MergedCellRange
import pandas as pd
from pydantic import BaseModel


class ScheduleEntry(BaseModel):
    """Parsed schedule entry"""
    group_name: str
    day_of_week: int  # 0=Monday, 6=Sunday
    pair_number: int  # 1, 2, 3 (I, II, III)
    week_type: str  # ODD or EVEN
    subject: str | None = None
    teacher: str | None = None
    room: str | None = None
    time: str | None = None


class ScheduleParser:
    """Parser for Excel schedule files with merged cells"""
    
    def __init__(self, file_content: bytes):
        """
        Initialize parser with Excel file content.
        
        Args:
            file_content: Raw bytes of the Excel file
        """
        self.file_content = file_content
        self.workbook = None
        self.sheet = None
        self.merged_ranges: list[MergedCellRange] = []
        self.unmerged_data: dict[str, Any] = {}
    
    def _load_workbook(self):
        """Load workbook with data_only=True to get calculated values"""
        if self.workbook is None:
            self.workbook = load_workbook(
                io.BytesIO(self.file_content),
                data_only=True
            )
            # Use first sheet by default
            self.sheet = self.workbook.active
            self.merged_ranges = list(self.sheet.merged_cells.ranges)
    
    def _unmerge_cells(self):
        """
        Unmerge all merged cells by replicating the top-left anchor value
        into all constituent cells.
        """
        self._load_workbook()
        
        # Create a copy of cell values
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    coord = f"{cell.column_letter}{cell.row}"
                    self.unmerged_data[coord] = cell.value
        
        # Process merged ranges
        for merged_range in self.merged_ranges:
            # Get the top-left cell (anchor)
            top_left_cell = self.sheet[merged_range.min_col][merged_range.min_row - 1]
            anchor_value = top_left_cell.value
            
            if anchor_value is None:
                continue
            
            # Replicate anchor value to all cells in the merged range
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = f"{self.sheet.cell(row=row_idx, column=col_idx).column_letter}{row_idx}"
                    self.unmerged_data[coord] = anchor_value
    
    def _get_cell_value(self, row: int, col: int) -> Any:
        """
        Get cell value, checking unmerged data first.
        
        Args:
            row: Row number (1-indexed)
            col: Column number (1-indexed)
        
        Returns:
            Cell value
        """
        coord = f"{self.sheet.cell(row=row, column=col).column_letter}{row}"
        
        # Check unmerged data first
        if coord in self.unmerged_data:
            return self.unmerged_data[coord]
        
        # Fallback to direct cell access
        cell = self.sheet.cell(row=row, column=col)
        return cell.value
    
    def _detect_week_split(self, text: str) -> tuple[str, str]:
        """
        Detect if a cell contains split week information (numerator/denominator).
        
        Args:
            text: Cell text content
        
        Returns:
            Tuple of (odd_week_text, even_week_text)
        """
        if not text or not isinstance(text, str):
            return (text, text)
        
        # Check for / delimiter
        if "/" in text:
            parts = text.split("/", 1)
            return (parts[0].strip(), parts[1].strip())
        
        # Check for newline delimiter
        if "\n" in text:
            parts = text.split("\n", 1)
            return (parts[0].strip(), parts[1].strip())
        
        # No split detected, same value for both weeks
        return (text, text)
    
    def _find_header_row(self) -> int | None:
        """
        Find the header row containing column names.
        
        Returns:
            Row number (1-indexed) or None
        """
        self._load_workbook()
        
        # Common header keywords
        header_keywords = ["day", "kun", "pair", "par", "time", "vaqt", "subject", "fan", "teacher", "o'qituvchi", "room", "xona"]
        
        for row_idx in range(1, min(20, self.sheet.max_row + 1)):  # Check first 20 rows
            row_values = []
            for col_idx in range(1, min(10, self.sheet.max_column + 1)):  # Check first 10 columns
                value = self._get_cell_value(row_idx, col_idx)
                if value:
                    row_values.append(str(value).lower())
            
            row_text = " ".join(row_values)
            if any(keyword in row_text for keyword in header_keywords):
                return row_idx
        
        return 1  # Default to first row
    
    def _parse_column_indices(self, header_row: int) -> dict[str, int]:
        """
        Parse column indices from header row.
        
        Args:
            header_row: Header row number
        
        Returns:
            Dictionary mapping column names to column indices
        """
        indices = {}
        
        for col_idx in range(1, self.sheet.max_column + 1):
            value = self._get_cell_value(header_row, col_idx)
            if not value:
                continue
            
            value_lower = str(value).lower()
            
            # Day column
            if "day" in value_lower or "kun" in value_lower:
                indices["day"] = col_idx
            
            # Pair column
            elif "pair" in value_lower or "par" in value_lower:
                indices["pair"] = col_idx
            
            # Time column
            elif "time" in value_lower or "vaqt" in value_lower:
                indices["time"] = col_idx
            
            # Subject column
            elif "subject" in value_lower or "fan" in value_lower:
                indices["subject"] = col_idx
            
            # Teacher column
            elif "teacher" in value_lower or "o'qituvchi" in value_lower or "oqtuvchi" in value_lower:
                indices["teacher"] = col_idx
            
            # Room column
            elif "room" in value_lower or "xona" in value_lower:
                indices["room"] = col_idx
            
            # Group column (usually first column)
            elif "group" in value_lower or "guruh" in value_lower:
                indices["group"] = col_idx
        
        return indices
    
    def _parse_day_of_week(self, day_text: str) -> int | None:
        """
        Parse day of week from text.
        
        Args:
            day_text: Day text (e.g., "Monday", "Dushanba", "1")
        
        Returns:
            Day of week (0=Monday, 6=Sunday) or None
        """
        if not day_text:
            return None
        
        day_text_lower = str(day_text).lower().strip()
        
        # English days
        days_en = {
            "monday": 0, "mon": 0,
            "tuesday": 1, "tue": 1,
            "wednesday": 2, "wed": 2,
            "thursday": 3, "thu": 3,
            "friday": 4, "fri": 4,
            "saturday": 5, "sat": 5,
            "sunday": 6, "sun": 6,
        }
        
        # Uzbek days
        days_uz = {
            "dushanba": 0, "du": 0,
            "seshanba": 1, "se": 1,
            "chorshanba": 2, "chor": 2,
            "payshanba": 3, "pay": 3,
            "juma": 4, "ju": 4,
            "shanba": 5, "sh": 5,
            "yakshanba": 6, "yak": 6,
        }
        
        if day_text_lower in days_en:
            return days_en[day_text_lower]
        if day_text_lower in days_uz:
            return days_uz[day_text_lower]
        
        # Try numeric (1-7, where 1=Monday)
        try:
            day_num = int(day_text_lower)
            if 1 <= day_num <= 7:
                return day_num - 1
        except ValueError:
            pass
        
        return None
    
    def _parse_pair_number(self, pair_text: str) -> int | None:
        """
        Parse pair number from text.
        
        Args:
            pair_text: Pair text (e.g., "I", "II", "III", "1", "2", "3")
        
        Returns:
            Pair number (1, 2, or 3) or None
        """
        if not pair_text:
            return None
        
        pair_text_upper = str(pair_text).upper().strip()
        
        # Roman numerals
        roman_to_num = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5}
        if pair_text_upper in roman_to_num:
            return roman_to_num[pair_text_upper]
        
        # Numeric
        try:
            pair_num = int(pair_text_upper)
            if 1 <= pair_num <= 5:
                return pair_num
        except ValueError:
            pass
        
        return None
    
    def parse(self) -> list[ScheduleEntry]:
        """
        Parse the Excel file and extract schedule entries.
        
        Returns:
            List of ScheduleEntry objects
        """
        self._load_workbook()
        self._unmerge_cells()
        
        header_row = self._find_header_row()
        column_indices = self._parse_column_indices(header_row)
        
        entries = []
        current_group = None
        
        # Start parsing from row after header
        for row_idx in range(header_row + 1, self.sheet.max_row + 1):
            # Get group name (usually first column)
            if "group" in column_indices:
                group_value = self._get_cell_value(row_idx, column_indices["group"])
                if group_value:
                    current_group = str(group_value).strip()
            
            # Skip if no group identified
            if not current_group:
                continue
            
            # Extract schedule data
            day_value = self._get_cell_value(row_idx, column_indices.get("day", 1)) if "day" in column_indices else None
            pair_value = self._get_cell_value(row_idx, column_indices.get("pair", 2)) if "pair" in column_indices else None
            time_value = self._get_cell_value(row_idx, column_indices.get("time", 3)) if "time" in column_indices else None
            subject_value = self._get_cell_value(row_idx, column_indices.get("subject", 4)) if "subject" in column_indices else None
            teacher_value = self._get_cell_value(row_idx, column_indices.get("teacher", 5)) if "teacher" in column_indices else None
            room_value = self._get_cell_value(row_idx, column_indices.get("room", 6)) if "room" in column_indices else None
            
            # Parse day and pair
            day_of_week = self._parse_day_of_week(day_value) if day_value else None
            pair_number = self._parse_pair_number(pair_value) if pair_value else None
            
            if day_of_week is None or pair_number is None:
                continue
            
            # Handle week splits
            subject_odd, subject_even = self._detect_week_split(subject_value) if subject_value else (None, None)
            teacher_odd, teacher_even = self._detect_week_split(teacher_value) if teacher_value else (None, None)
            room_odd, room_even = self._detect_week_split(room_value) if room_value else (None, None)
            
            # Create entries for both weeks
            entries.append(ScheduleEntry(
                group_name=current_group,
                day_of_week=day_of_week,
                pair_number=pair_number,
                week_type="ODD",
                subject=subject_odd,
                teacher=teacher_odd,
                room=room_odd,
                time=str(time_value) if time_value else None,
            ))
            
            entries.append(ScheduleEntry(
                group_name=current_group,
                day_of_week=day_of_week,
                pair_number=pair_number,
                week_type="EVEN",
                subject=subject_even,
                teacher=teacher_even,
                room=room_even,
                time=str(time_value) if time_value else None,
            ))
        
        return entries
